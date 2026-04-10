#!/usr/bin/env python3
"""
FP360 Market Intelligence Runner
- Pulls RSS feeds + optional SEC EDGAR submissions
- Dedupes and appends to Raw_Feed + Signal_Log in the Excel workbook
- Applies simple heuristic scoring + posture assignment
Notes:
- This script is designed to run on YOUR machine (it needs internet).
- Obey SEC "Fair Access" guidance and include a real User-Agent with contact info.
SEC docs: https://www.sec.gov/search-filings/edgar-application-programming-interfaces
MINING.com RSS reference: https://www.mining.com/mining-dot-com-rss-feeds/
"""

import hashlib
import re
import time
import datetime as dt
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook

try:
    import feedparser  # pip install feedparser
except Exception:
    feedparser = None


WORKBOOK_PATH = "FP360_US_Market_Intelligence_Template.xlsx"  # change if needed

# --- Heuristic dictionaries (edit to match FP360 focus) ---
SEGMENT_RULES = [
    ("EPCM", r"\b(EPCM|engineering|procurement\s+and\s+construction|Fluor|Jacobs|Worley)\b"),
    ("Contractor", r"\b(contractor|mining\s+services|earthworks|drilling|blasting)\b"),
    ("OEM", r"\b(Caterpillar|Komatsu|Sandvik|Epiroc|Hitachi|OEM)\b"),
    ("Supplier", r"\b(MRO|spares|consumables|parts|supplier)\b"),
    ("Startup/Platform", r"\b(startup|platform|software|AI|digital|automation)\b"),
]
CATEGORY_RULES = [
    ("Growth", r"\b(expand|expansion|ramp[-\s]?up|scale|increase production|growth)\b"),
    ("Entry", r"\b(enters? the U\.S\.|new entrant|launch|start operations|new project)\b"),
    ("Change", r"\b(ERP|implementation|restructure|reorg|operating model|transformation)\b"),
    ("Risk", r"\b(delay|shortage|disruption|cost overrun|missed|halt|suspend|bankrupt)\b"),
    ("Transition", r"\b(energy transition|battery|lithium|REE|rare earth|nickel|cobalt)\b"),
]

PRESSURE_POINT_RULES = [
    ("Procurement", r"\b(procurement|sourcing|tender|supplier|contract award)\b"),
    ("Inventory", r"\b(inventory|spares|MRO|stockout|overstock|warehouse)\b"),
    ("Planning", r"\b(plan|schedule|ramp|forecast|capacity)\b"),
    ("Supply Chain", r"\b(lead time|logistics|shipping|transport|port|rail)\b"),
    ("Operations", r"\b(operations|commission|startup|production|throughput)\b"),
]

US_KEYWORDS = re.compile(r"\b(United States|U\.S\.|USA|Nevada|Arizona|Utah|Texas|Wyoming|Alaska|Idaho|New Mexico|Colorado|Montana|Minnesota|Michigan)\b", re.I)

def now_iso() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")

def dedupe_key(link: str, title: str) -> str:
    h = hashlib.sha1((link.strip() + "|" + title.strip()).encode("utf-8")).hexdigest()
    return h[:16]

def classify(text: str, rules: List[Tuple[str, str]], default: str) -> str:
    for label, pattern in rules:
        if re.search(pattern, text, flags=re.I):
            return label
    return default

def score_dimensions(text: str) -> Tuple[int,int,int,int]:
    """
    Returns (impact 0-30, time_to_strain 0-25, fit 0-25, access 0-20)
    Heuristic only. You can replace this with your own logic or an LLM classifier.
    """
    t = text.lower()

    # Operational impact
    impact = 5
    if re.search(r"\b(expansion|ramp|restart|new mine|processing plant|capex)\b", t): impact += 10
    if re.search(r"\b(erp|transformation|operating model|restructure)\b", t): impact += 8
    if re.search(r"\b(merger|acquisition|divest|spin)\b", t): impact += 7
    impact = min(30, impact)

    # Time-to-strain
    tts = 6
    if re.search(r"\b(immediate|this quarter|2026|next month|short-term)\b", t): tts += 10
    if re.search(r"\b(startup|commission|ramp)\b", t): tts += 6
    tts = min(25, tts)

    # FP360 fit
    fit = 6
    if re.search(r"\b(procurement|sourcing|supplier|inventory|planning|lead time|logistics)\b", t): fit += 12
    if re.search(r"\b(spares|MRO|maintenance)\b", t): fit += 5
    fit = min(25, fit)

    # Commercial access
    access = 5
    if re.search(r"\b(appoints|named|joins as|new CEO|new COO|new CFO|VP operations|head of supply chain)\b", t): access += 8
    if re.search(r"\b(funding|financing|raise|IPO|S-1|investment)\b", t): access += 5
    if re.search(r"\b(transformation|initiative|program)\b", t): access += 3
    access = min(20, access)

    return impact, tts, fit, access

def posture(total_score: int) -> str:
    if total_score >= 80: return "Proactive Outreach"
    if total_score >= 65: return "Target Account"
    if total_score >= 45: return "Prepare POV"
    return "Monitor"

def confidence(total_score: int) -> str:
    if total_score >= 80: return "High"
    if total_score >= 65: return "Medium"
    return "Low"

def fetch_rss(url: str, user_agent: str) -> List[Dict]:
    if feedparser is None:
        raise RuntimeError("feedparser not installed. Run: pip install feedparser")
    fp = feedparser.parse(url, request_headers={"User-Agent": user_agent})
    items = []
    for e in fp.entries:
        items.append({
            "source": url,
            "title": getattr(e, "title", "") or "",
            "link": getattr(e, "link", "") or "",
            "published": getattr(e, "published", "") or getattr(e, "updated", "") or "",
            "summary": re.sub(r"\s+", " ", re.sub("<.*?>", "", getattr(e, "summary", "") or ""))[:500],
            "tags": ",".join([t.term for t in getattr(e, "tags", [])]) if hasattr(e, "tags") else ""
        })
    return items

def workbook_has_key(ws, key: str) -> bool:
    # assumes Dedupe_Key is last column in both Raw_Feed and Signal_Log
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[-1] == key:
            return True
    return False

def main():
    wb = load_workbook(WORKBOOK_PATH)
    ws_config = wb["Config"]
    ws_raw = wb["Raw_Feed"]
    ws_log = wb["Signal_Log"]

    # Read config values
    cfg = {}
    for r in ws_config.iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            cfg[str(r[0]).strip()] = str(r[1]).strip() if r[1] is not None else ""

    user_agent = cfg.get("USER_AGENT", "FP360MarketIntel/1.0 (contact: you@example.com)")
    rss_default = cfg.get("MININGCOM_RSS_DEFAULT", "https://www.mining.com/feed/")

    feeds = [rss_default]  # extend this list with more feeds as you like

    ingested_at = now_iso()
    new_count = 0

    for feed in feeds:
        items = fetch_rss(feed, user_agent=user_agent)
        for it in items:
            text_blob = f"{it['title']} {it['summary']}"
            us_flag = "Yes" if US_KEYWORDS.search(text_blob) else "Maybe"

            key = dedupe_key(it["link"], it["title"])
            if workbook_has_key(ws_raw, key) or workbook_has_key(ws_log, key):
                continue

            # Raw_Feed append
            ws_raw.append([
                ingested_at,
                "RSS",
                it["title"],
                it["published"],
                it["link"],
                it["summary"],
                it["tags"],
                us_flag,
                key
            ])

            # Signal_Log append (only if US relevant at least Maybe; tune to strict "Yes" if desired)
            if us_flag in ("Yes","Maybe"):
                seg = classify(text_blob, SEGMENT_RULES, default="Operator")
                cat = classify(text_blob, CATEGORY_RULES, default="Change")
                pp = classify(text_blob, PRESSURE_POINT_RULES, default="Supply Chain")

                impact, tts, fit, access = score_dimensions(text_blob)
                total = impact + tts + fit + access

                ws_log.append([
                    dt.date.today().isoformat(),     # Date_Detected
                    "",                               # Company (optional: add extractor)
                    seg,                              # Segment
                    it["title"][:280],                # Signal_Summary
                    cat,                              # Category
                    "MINING.com RSS",                 # Source
                    impact, tts, fit, access,
                    None,                             # Total_Score formula in sheet will compute
                    confidence(total),                # Confidence
                    pp,                               # Likely_Pressure_Point
                    posture(total),                   # BD_Posture
                    "",                               # Owner
                    "",                               # Review_Week
                    it["summary"][:450],              # Notes
                    it["link"],                       # Link
                    key                               # Dedupe_Key
                ])
                new_count += 1

        time.sleep(1.0)  # be polite to sources

    wb.save(WORKBOOK_PATH)
    print(f"Done. Added {new_count} new signal(s).")

if __name__ == "__main__":
    main()
